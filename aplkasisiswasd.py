import os
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf as matplotlib_pdf
from matplotlib.backends.backend_pdf import PdfPages
from pandas.plotting import table
from sklearn.cluster import KMeans
from openpyxl.cell.cell import MergedCell
from sklearn.preprocessing import MinMaxScaler
from streamlit_option_menu import option_menu
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import datetime
import xlsxwriter
from io import BytesIO
import datetime
from openpyxl.worksheet.page import PageMargins
import requests
from PIL import Image as PILImage
from io import BytesIO as PILBytesIO
from openpyxl.utils import get_column_letter
import webbrowser
import time

def open_browser():
    time.sleep(1)  # Tunggu sedikit untuk memastikan Streamlit server sudah berjalan
    webbrowser.open_new("http://localhost:8501")


# User credentials
usern = "admin"
passw = "admin"
logo_url = r'C:\Users\AHMED\Aplikasi Kelas Unggulan\Bogor.png'
logo_path = r'C:\Users\AHMED\Aplikasi Kelas Unggulan\Bogor.png'   

# Function to get current date in Indonesian format
def get_indonesian_date():
    months = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    now = datetime.datetime.now()
    return f"{now.day}, {months[now.month]} {now.year}"

# Function to get current day in Indonesian format
def get_indonesian_day():
    days = {
        0: "Senin", 1: "Selasa", 2: "Rabu", 3: "Kamis",
        4: "Jumat", 5: "Sabtu", 6: "Minggu"
    }
    # Get the current day as a number (0=Monday, 1=Tuesday, ..., 6=Sunday)
    day_number = datetime.now().weekday()
    # Return the corresponding day in Indonesian
    return days[day_number]

# Custom CSS for Streamlit
st.markdown("""
    <style>
    body {
        background-color: #f0f0f0; /* Light grey background for the whole page */
    }
    .stButton>button {
        background-color: #003366; /* Dark blue button */
        color: white;
        border: none;
        padding: 10px 24px;
        text-align: center;
        text-decoration: none;
        display: inline-block;
        font-size: 16px;
        margin: 4px 2px;
        cursor: pointer;
        border-radius: 12px;
    }
    .stTextInput>div>div>input {
        border: 2px solid #003366; /* Dark blue border */
        border-radius: 12px;
    }
    .stTextInput>div>div>input:focus {
        border-color: #000000; /* Darker blue for focus */
    }
    .stAlert {
        border-radius: 12px;
    }
    .stMarkdown {
        color: #000000; /* Black text */
    }
    .stSidebar {
        background-color: #003366; /* Dark blue sidebar */
        color: white;
    }
    .stSidebar .stButton>button {
        background-color: #001a33; /* Darker blue button in sidebar */
    }
    .stSidebar .stButton>button:hover {
        background-color: #002244; /* Even darker blue for hover */
    }
    .stSidebar .stMarkdown {
        color: white; /* White text in sidebar */
    }
    .stSidebar .stOptionMenu {
        background-color: #003366; /* Dark blue background for options */
    }
    .stSidebar .stOptionMenu .nav-link {
        color: black; /* Black text for menu items */
    }
    .stSidebar .stOptionMenu .nav-link:hover {
        background-color: white; /* White background for hover */
        color: #003366; /* Dark blue text for hover */
    }
    .stSidebar .stOptionMenu .nav-link-selected {
        background-color: #003366; /* Dark blue background for selected menu item */
        color: white; /* White text for selected menu item */
    }
    .stSidebar .stOptionMenu .nav-link .icon {
        color: white; /* White icon color */
    }
    .stSidebar .stOptionMenu .nav-link:hover .icon {
        color: #003366; /* Dark blue icon color on hover */
    }
    /* Custom CSS to position the date at the bottom-right corner of the data frame */
    .date-container {
        position: absolute;
        bottom: 1.5px;
        right: 10px;
        font-size: 14px;
        color: #000000;
    }
    .plt.figure .date_info .principal_name {
        text-decoration: underline;
    }
    </style>
""", unsafe_allow_html=True)

# Login function
def login():
    st.markdown('<div class="login-container" style="text-align: center;">', unsafe_allow_html=True)
    st.image(logo_url, width=130)  # Logo Login
    st.title("Halaman Login Aplikasi K-means Clustering Pembagian Kelas Unggulan SDN IPK Ciriung 01")
    st.markdown("## Silahkan Masuk")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login"):
        if username == usern and password == passw:
            st.session_state['logged_in'] = True
            st.success("Anda berhasil login")
            st.experimental_rerun()  # Refresh halaman setelah login
        else:
            st.error("Username dan password tidak cocok")
    st.markdown('</div>', unsafe_allow_html=True)

# K-Means Analysis Function
def kmeans_analysis(df, n_clusters):
    # Menghitung rata-rata nilai UTS dan UAS serta rata-rata gabungan
    df['Rata-Rata UTS'] = df[['UTS Matematika', 'UTS Indonesia', 'UTS IPA', 'UTS Inggris']].mean(axis=1)
    df['Rata-Rata UAS'] = df[['UAS Matematika', 'UAS Indonesia', 'UAS IPA', 'UAS Inggris']].mean(axis=1)
    df['Keseluruhan'] = (df['Rata-Rata UTS'] + df['Rata-Rata UAS']) / 2
    
    # Menyiapkan data untuk clustering
    x_train = df[['Rata-Rata UTS', 'Rata-Rata UAS', 'Keseluruhan']].values
    scaler = MinMaxScaler()
    x_train_scaled = scaler.fit_transform(x_train)
    x_train_scaled = x_train_scaled.astype(np.float64)
    
    # Melakukan clustering
    kmean = KMeans(n_clusters=n_clusters)
    kmean.fit(x_train_scaled)
    y_cluster = kmean.predict(x_train_scaled)
    df['Cluster'] = y_cluster
    
    # Menambahkan nama cluster berdasarkan rata-rata gabungan
    cluster_names = ['C' if (val >= 70 and val <= 79) else
                     ('B' if (val >= 80 and val <= 85) else 'Unggulan')
                     for val in df['Keseluruhan']]
    df['Kelas'] = cluster_names
    
    # Plot distribusi siswa pada pembagian kelas
    cluster_counts = df['Kelas'].value_counts()
    fig_bar = plt.figure(figsize=(10, 6))
    cluster_counts.plot(kind='bar', color='steelblue')
    plt.xlabel('Kelas')
    plt.ylabel('Jumlah Siswa')
    plt.title('Distribusi Siswa Pada Pembagian Kelas')
    plt.xticks(rotation=0)
    plt.tight_layout()
    
    # Plot sebaran nilai UTS dan UAS siswa
    fig_scatter = plt.figure(figsize=(10, 6))
    plt.scatter(df['Rata-Rata UTS'], df['Rata-Rata UAS'], c=df['Cluster'], cmap='viridis', alpha=0.6)
    plt.xlabel('Rata-rata Nilai UTS')
    plt.ylabel('Rata-rata Nilai UAS')
    plt.title('Plot Sebaran Nilai UTS dan UAS Siswa')
    plt.colorbar(label='Centroid Cluster')
    plt.tight_layout()
    
    return df, fig_bar, fig_scatter

def save_descriptive_stats_to_excel(df, principal_name, logo_path):
    # Create a BytesIO buffer
    output = BytesIO()

    # Use Pandas ExcelWriter with openpyxl engine
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.describe().to_excel(writer, sheet_name='Descriptive Stats', startrow=4)  # Start from row 4
        workbook = writer.book
        worksheet = writer.sheets['Descriptive Stats']

        # Add logo in header
        logo = Image(logo_path)
        logo.height = 100  # Adjust the size of the logo
        logo.width = 100
        worksheet.add_image(logo, 'A1')

        # Merge cells for title, school name, and address
        worksheet.merge_cells('B1:O1')
        worksheet.merge_cells('B2:O2')
        worksheet.merge_cells('B3:O3')

        # Adding title, school name, and address to header
        worksheet['B1'] = 'TABEL DESKRIPSI DATASET'
        worksheet['B2'] = 'SDN IPK CIRIUNG 01'
        worksheet['B3'] = 'Jl. Mayor Oking Jaya Atmaja CIRIUNG, Kec. Cibinong, Kab. Bogor, Jawa Barat, 16918'

        # Setting center alignment and font size
        for cell in ['B1', 'B2', 'B3']:
            worksheet[cell].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[cell].font = Font(bold=True, size=16)  # Bold and font size 16 for title and school name
        worksheet['B3'].font = Font(bold=False, size=12)  # Regular and font size 12 for address
        

        # Define border style
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Auto adjust column width
        for col in worksheet.columns:
            max_length = 0
            column_index = col[0].column  # Get the column index
            column_letter = get_column_letter(column_index)  # Convert to letter
            for cell in col:
                if not isinstance(cell, MergedCell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = min(max_length + 1.5, 20)  # Set maximum width to 20
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply border only to the dataset cells (starting from row 5)
        data_start_row = 5
        data_end_row = data_start_row + len(df.describe())
        data_end_col = len(df.columns) + 1  # +1 to adjust for the index column

        for row in worksheet.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=1, max_col=data_end_col):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.border = border_style

        # Create footer based on image layout
        row_count = data_end_row + 2  # Adjust starting row as needed
        col_count = len(df.describe().columns) 

        footer_lines = [
            f'Cibinong, {get_indonesian_day()}, {get_indonesian_date()}',
            f'Mengetahui,',
            f'Kepala Sekolah',
            '',
            '',
            '',
            f'{principal_name}'
        ]

        # Write footer content with appropriate formatting and cell merging
        for idx, line in enumerate(footer_lines):
            cell = worksheet.cell(row=row_count + idx, column=col_count)
            cell.value = line
            if line == principal_name:
                cell.font = Font(underline='single')  # Add underline to principal's name

        # Align footer text based on position (assuming 'Mengetahui' is left-aligned, others are right-aligned)
        for idx, line in enumerate(footer_lines):
            cell = worksheet.cell(row=row_count + idx, column=col_count)
            cell.value = line
            if idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='top')

        # Merge cells for footer content (assuming 'Mengetahui' on a separate row)
        if len(footer_lines) > 3:
            worksheet.merge_cells(start_row=row_count + 3, start_column=col_count,
                                 end_row=row_count + len(footer_lines) - 3, end_column=col_count)

    # Return the Excel file as bytes
    return output.getvalue()

# Downoad Pemagian Kelas Unggulan Pada Excel
from datetime import datetime



def get_indonesian_day():
    days = {
        0: "Senin", 1: "Selasa", 2: "Rabu", 3: "Kamis",
        4: "Jumat", 5: "Sabtu", 6: "Minggu"
    }
    # Get the current day as a number (0=Monday, 1=Tuesday, ..., 6=Sunday)
    day_number = datetime.now().weekday()
    # Return the corresponding day in Indonesian
    return days[day_number]

def get_indonesian_date():
    months = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    now = datetime.now()
    return f"{now.day} {months[now.month]} {now.year}"

def to_excel_with_footer(df, principal_name, logo_path):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=5)  # startrow=5 to start from row 6
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Add logo in header
        logo = Image(logo_path)
        logo.height = 100  # Adjust the size of the logo
        logo.width = 100
        worksheet.add_image(logo, 'A1')

        # Merge cells for title, school name, and address
        worksheet.merge_cells('B1:O1')
        worksheet.merge_cells('B2:O2')
        worksheet.merge_cells('B3:O3')

        # Adding title, school name, and address to header
        worksheet['B1'] = 'TABEL PEMBAGIAN KELAS'
        worksheet['B2'] = 'SDN IPK CIRIUNG 01'
        worksheet['B3'] = 'Jl. Mayor Oking Jaya Atmaja CIRIUNG, Kec. Cibinong, Kab. Bogor, Jawa Barat, 16918'

        # Setting center alignment and font size
        for cell in ['B1', 'B2', 'B3']:
            worksheet[cell].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[cell].font = Font(bold=True, size=16)  # Bold and font size 16 for title and school name
        worksheet['B3'].font = Font(bold=False, size=12)  # Regular and font size 12 for address
        
        # Adjust column widths to fit data
        for col in worksheet.columns:
            max_length = 0
            column_index = col[0].column  # Get the column index
            column_letter = get_column_letter(column_index)  # Convert to letter
            for cell in col:
                if not isinstance(cell, MergedCell):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = max_length + 0.4
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Set borders for all cells in the table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply borders and center alignment for all cells in the table
        for row in worksheet.iter_rows(min_row=6, max_row=len(df) + 6, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create footer based on image layout
        row_count = len(df) + 8  # Adjust starting row as needed
        col_count = len(df.columns)

        footer_lines = [
            f'Cibinong, {get_indonesian_day()}, {get_indonesian_date()}',  # Capture day and date using get_indonesian_day() and get_indonesian_date() functions
            f'Mengetahui,',
            f'Kepala Sekolah',
            '',
            '',
            '',
            f'{principal_name}'
        ]

        # Write footer content with appropriate formatting and cell merging
        for idx, line in enumerate(footer_lines):
            cell = worksheet.cell(row=row_count + idx, column=col_count)
            cell.value = line
            if line == principal_name:
                cell.font = Font(underline='single')  # Add underline to principal's name

        # Align footer text based on position (assuming 'Mengetahui' is left-aligned, others are right-aligned)
        for idx, line in enumerate(footer_lines):
            cell = worksheet.cell(row=row_count + idx, column=col_count)
            cell.value = line
            if idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='top')
        
        # Merge cells for footer content (assuming 'Mengetahui' on a separate row)
        if len(footer_lines) > 3:
            worksheet.merge_cells(start_row=row_count + 3, start_column=col_count,
                                 end_row=row_count + len(footer_lines) - 3, end_column=col_count)

    processed_data = output.getvalue()
    return processed_data



#Laporan Centroid Dan Distribusi siswa
def add_text_and_logo_to_plot(fig, principal_name, logo_path):
    plt.figure(fig.number)

    # Load and add logo
    logo = PILImage.open(logo_path)
    logo = logo.resize((100, 100))  # Resize logo as needed
    fig.figimage(logo, xo=80, yo=fig.get_figheight()*fig.dpi - 150, zorder=1)  # Adjust position and size

    # Format text with line breaks for footer
    date_info = (
        f'Cibinong, {get_indonesian_day()}, {get_indonesian_date()}\n'
        'Mengetahui,                                  \n'
        'Kepala Sekolah                             \n\n\n\n\n\n'  # Empty lines for spacing
        f'{principal_name}               '
    )
    
    # Add school name and address at the top
    school_info = (
        'SDN IPK CIRIUNG 01\n'
        )
    
    alamat_sklh = (
        'Jl. Mayor Oking Jaya Atmaja CIRIUNG, Kec. Cibinong, Kab. Bogor, Jawa Barat, 16918'
    )
    
    # Adjust layout to make space at the top and bottom
    plt.subplots_adjust(top=0.75, bottom=0.40)  # Top and bottom adjustments for space

    # Add text with figtext() for footer
    plt.figtext(0.95, 0.05, date_info, horizontalalignment='right', verticalalignment='bottom',
                fontsize=10, color='black',
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='none'))

    # Add text with figtext() for school information
    plt.figtext(0.5, 0.95, school_info, horizontalalignment='center', verticalalignment='top',
                fontsize=14, color='black',
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='none'))
    
    # Add text with figtext() for school information
    plt.figtext(0.5, 0.90, alamat_sklh, horizontalalignment='center', verticalalignment='top',
                fontsize=9, color='black',
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='none'))

def fig_to_pdf(fig, principal_name, logo_path):
    add_text_and_logo_to_plot(fig, principal_name, logo_path)
    output = BytesIO()
    fig.savefig(output, format='pdf', bbox_inches='tight')  # Ensure to include tight bounding box
    output.seek(0)
    return output


# Main app logic
if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

if not st.session_state['logged_in']:
    login()
else:
    with st.sidebar:
        selected = option_menu('',
                               ['Tentang Sekolah', 'Tentang K-Means Cluster', 'Hitung Pembagian Kelas Unggulan'],
                               icons=['info-square', 'book', 'clipboard2-data'], menu_icon="menu-button-wide", default_index=0,
                               styles={
                                   "container": {"padding": "5px", "background-color": "#ffffff"},
                                   "icon": {"color": "#00000", "font-size": "25px"}, 
                                   "nav-link": {"font-size": "20px", "text-align": "left", "margin": "0px"},
                                   "nav-link-hover": {"background-color": "white", "color": "#003366"},
                                   "nav-link-selected": {"background-color": "#003366", "color": "white"},
                               })
        if st.button('Logout'):
            st.session_state['logged_in'] = False
            st.experimental_rerun()      

    if selected == 'Hitung Pembagian Kelas Unggulan':
        st.title('Menghitung Pembagian Kelas Unggulan SDN IPK Ciriung 01')
        uploaded_file = st.file_uploader("Pilih file Excel", type=["xlsx"])
        kriteria = r'C:\Users\AHMED\Aplikasi Kelas Unggulan\kriteria.png'
        st.image(kriteria, caption='Sesuaikan Format Seperti Pada Gambar', use_column_width=True)
        if uploaded_file is not None:
            df = pd.read_excel(uploaded_file)
            n_clusters = st.slider('Jumlah Cluster (K)', min_value=1, max_value=3, value=3)
            if st.button("Hitung Algoritma K-means Clustering"):
                df, fig_bar, fig_scatter = kmeans_analysis(df, n_clusters)
                current_date = get_indonesian_date()
                principal_name = "HJ.YULIAWATI, S.Pd.M.Pd"
                
                st.header("Deskripsi Statistik Dataset")
                st.write(df.describe())

                st.download_button(
                    label="Download Statistik Deskriptif sebagai Excel",
                    data=save_descriptive_stats_to_excel(df, principal_name, logo_path),
                    file_name='statistik_deskriptif.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                st.header("Pembagian Kelas Siswa")
                st.dataframe(df)

                st.download_button(
                    label="Download File Excel",
                    data= to_excel_with_footer(df, principal_name, logo_path),
                    file_name='pembagian_kelas_unggulan.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

                st.header("Memvisualisasikan K-Means Clustering (Bar Plot)")
                st.pyplot(fig_bar)
                st.download_button(
                    label="Download Bar Plot as PDF",
                    data=fig_to_pdf(fig_bar, principal_name, logo_path),
                    file_name='bar_plot.pdf',
                    mime='application/pdf'
                )

                st.header("Plot Sebaran Total Nilai Siswa Menggunakan Centroid")
                st.pyplot(fig_scatter)
                st.download_button(
                    label="Download Scatter Plot as PDF",
                    data=fig_to_pdf(fig_scatter, principal_name, logo_path),
                    file_name='scatter_plot.pdf',
                    mime='application/pdf'
                )
        else:
            st.info("Silahkan Masukan File dengan Format Excel")

    if selected == 'Tentang K-Means Cluster':
        st.title('Tentang K-Means Cluster')
        current_date = get_indonesian_date()
        st.write(f"""
        ## Apa itu K-Means Clustering?
        K-Means Clustering adalah salah satu algoritma clustering yang paling populer dan sering digunakan dalam analisis data. Algoritma ini bertujuan untuk membagi dataset ke dalam beberapa kelompok (cluster) yang berbeda, di mana data dalam setiap cluster memiliki karakteristik yang mirip satu sama lain dan berbeda dari data di cluster lain.

        ### Bagaimana K-Means Clustering Bekerja?
        Proses K-Means Clustering dapat dijelaskan dalam beberapa langkah sebagai berikut:
        1. **Menentukan Jumlah Cluster (K)**: Sebagai langkah awal, kita harus menentukan berapa jumlah cluster (K) yang diinginkan.
        2. **Menginisialisasi Centroid**: Pilih secara acak K titik dalam dataset sebagai titik pusat awal (centroid) untuk setiap cluster.
        3. **Mengalokasikan Setiap Titik ke Cluster Terdekat**: Setiap titik data diukur jaraknya ke setiap centroid dan dialokasikan ke cluster dengan centroid terdekat.
        4. **Mengupdate Centroid**: Setelah semua titik data dialokasikan ke cluster, hitung ulang posisi centroid sebagai rata-rata dari semua titik data dalam cluster tersebut.
        5. **Iterasi**: Ulangi langkah 3 dan 4 sampai centroid tidak lagi berubah secara signifikan atau sampai jumlah iterasi yang telah ditentukan tercapai.

        ### Kelebihan K-Means Clustering:
        - **Sederhana dan Cepat**: Algoritma ini relatif mudah diimplementasikan dan cepat dalam komputasi, terutama untuk dataset besar.
        - **Skalabilitas**: K-Means dapat dengan mudah diskalakan untuk menangani dataset yang sangat besar.

        ### Kelemahan K-Means Clustering:
        - **Pemilihan Jumlah Cluster (K)**: Algoritma ini memerlukan penentuan jumlah cluster (K) di awal, yang mungkin tidak selalu jelas.
        - **Sensitif Terhadap Inisialisasi Centroid**: Hasil clustering dapat bervariasi berdasarkan pemilihan centroid awal.
        - **Tidak Efektif untuk Bentuk Cluster Non-Bulat**: K-Means bekerja terbaik untuk cluster dengan bentuk bulat dan ukuran yang serupa, tetapi kurang efektif untuk cluster dengan bentuk yang kompleks atau ukuran yang sangat berbeda.

        ### Contoh Aplikasi K-Means Clustering:
        - **Segmentasi Pelanggan**: Mengelompokkan pelanggan berdasarkan pola pembelian atau preferensi mereka.
        - **Pengelompokan Dokumen**: Mengelompokkan dokumen teks berdasarkan kemiripan konten.
        - **Analisis Citra**: Mengelompokkan piksel dalam gambar untuk segmentasi citra.

        (Diperbarui pada {current_date})
        """)

    if selected == 'Tentang Sekolah':
        st.title("Tentang SDN IPK Ciriung 01")
        school_image_url = r'C:\Users\AHMED\Aplikasi Kelas Unggulan\sd.png'
        st.image(school_image_url, caption='Sekolah SDN IPK Ciriung 01', use_column_width=True)
        current_date = get_indonesian_date()
        st.write(f"""
        Sekolah SDN IPK Ciriung 01 terletak di Jl.mayor Oking Jaya Atmaja CIRIUNG,Kec. Cibinong, Kab. Bogor, Jawa Barat, 16918 dan dikenal sebagai salah satu sekolah dasar yang memberikan pendidikan berkualitas bagi siswa-siswinya. Sekolah ini memiliki berbagai fasilitas dan program pendidikan yang mendukung perkembangan akademis maupun non-akademis para siswa.

        ## Fasilitas Sekolah
        - **Laboratorium Komputer**: Dilengkapi dengan komputer dan koneksi internet untuk mendukung pembelajaran teknologi informasi.
        - **Perpustakaan**: Menyediakan berbagai buku dan bahan bacaan untuk menunjang kegiatan belajar mengajar.
        - **Lapangan Olahraga**: Fasilitas olahraga untuk mendukung kegiatan ekstrakurikuler dan kebugaran siswa.

        ## Program Unggulan
        - **Program Kelas Unggulan**: Membagi siswa berdasarkan prestasi akademik untuk memberikan perhatian dan pembelajaran yang lebih sesuai dengan kebutuhan mereka.
        - **Ekstrakurikuler**: Berbagai kegiatan ekstrakurikuler seperti pramuka, seni, dan olahraga untuk mengembangkan bakat dan minat siswa.

        ## Visi dan Misi
        **Visi**: Menjadi sekolah yang unggul dalam prestasi, berkarakter, dan berwawasan lingkungan.
        **Misi**:
        - Menyediakan pendidikan yang berkualitas dan berfokus pada pengembangan karakter siswa.
        - Mengembangkan potensi akademik dan non-akademik siswa melalui berbagai program dan kegiatan.
        - Menciptakan lingkungan belajar yang aman, nyaman, dan kondusif untuk proses pembelajaran.

        (Diperbarui pada {current_date})
        """)
